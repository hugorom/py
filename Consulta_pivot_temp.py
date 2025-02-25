# -*- coding: utf-8 -*-
"""
Created on Tue May 28 17:43:41 2024
Crea una matriz de datos dependiendo de la cadena solicitada
@author: hromanr
"""
import mariadb
import pandas as pd
import numpy as np
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import numbers
# Datos de conexión a la base de datos
config = {'user': 'disponibilidad', 'password': 'd_2024', 'host': '172.29.149.155', 'port': 3306, 'database': 'sihah'}
# Lista de cadenas que deseas utilizar en la consulta
estaciones = ['SOLVC', 'NOGVC', 'MALVC', 'C30486', 'C30120', 'TUPVC', 'PRLVC', 'PTXVC', 'LOMAGRANDE', 'PTUXPANGO', 'ORIZABAob', 
              'IFSANSEBASTIAN', 'PCRPB', 'LGRVC', 'X76737', 'C30309', 'C30299', 'C30274', 'C30212', 'C30100', 'C30042', 'C30004',
              'C21161', 'C21160', 'C21159', 'C21158', 'C21053', 'C21039', 'C21020', 'ORZVC']

try:
    conn = mariadb.connect(**config) # Conectar a la base de datos
    cursor = conn.cursor()
    
    # Crear la consulta SQL dinámica
    placeholders = ', '.join(['%s'] * len(estaciones))
    query = f"SELECT * FROM ddprecipitacio WHERE station IN ({placeholders}) and year(datee) >= 1940"
    cursor.execute(query, tuple(estaciones)) # Ejecutar la consulta
        
    resultados = cursor.fetchall() # Obtener los resultados
    columnas = [desc[0] for desc in cursor.description] # Obtener los nombres de las columnas
    
    # Crear un DataFrame de pandas con los resultados
    df = pd.DataFrame(resultados, columns=columnas)
    # Pivotar el DataFrame   Suponiendo que 'tu_columna' es la columna con los valores de la lista y     # quieres pivotar sobre 'tu_columna' y otra columna 'otra_columna'
    df_pivot = df.pivot(index='Datee', columns='Station', values='Valuee')
    ix = pd.date_range(name=df_pivot.index.name, start=('1940-01-01'), end=('2023-12-31'), freq='D')
    df_pivot = df_pivot.reindex(ix)
    
    df_analisis = df_pivot.copy()
    
    for col in df_pivot.columns:
        df_pivot.fillna({col:-999}, inplace=True)
    
    #print(df_pivot)
    ruta_carpeta = "C:\Z"  # Reemplaza con la ruta real
    nombre_archivo = "RB2_.xlsx"

    ruta_completa = os.path.join(ruta_carpeta, nombre_archivo)
    try:
        # Crear la carpeta
        os.makedirs(ruta_carpeta)
        print(f"Carpeta creada exitosamente en {ruta_carpeta}")
    except FileExistsError:
        print(f"La carpeta ya existe en {ruta_carpeta}")

        # Crear la ruta completa del archivo
    df_pivot.to_excel(ruta_completa)
    df_analisis['year'] = df_analisis.index.year
    
    resultados_agrupa = pd.DataFrame()
    resultados_med = pd.DataFrame()
    df_sums = pd.DataFrame()

    r_log = pd.DataFrame()
    r_log2 = pd.DataFrame()
    # Recorrer las columnas del DataFrame pivotado (excluyendo la columna 'year')
    for col in df_analisis.columns:
        if col != 'year':
            # Agrupar por año y contar la cantidad de datos anuales
            conteo_anual = df_analisis.groupby('year')[col].count().reset_index()
            conteo_anual.columns = ['year', f'{col}']
            # Unir los resultados en el DataFrame de resultados
            if resultados_agrupa.empty:
                resultados_agrupa = conteo_anual
            else:       
                resultados_agrupa = pd.merge(resultados_agrupa, conteo_anual, on='year', how='outer')
        
    for col in df_analisis.columns:
        if col != 'year':
            # Agrupar por año y contar la cantidad de datos anuales
            suma_ = df_analisis.groupby('year')[col].sum().reset_index()
            suma_.columns = ['year', f'{col}_s']
            conteo_anual = df_analisis.groupby('year')[col].count().reset_index()
            conteo_anual.columns = ['year', f'{col}_n']
            # Unir los resultados en el DataFrame de resultados
            if df_sums.empty:
                df_sums = suma_
            else:       
                df_sums = pd.merge(df_sums, suma_, on='year', how='outer')
            
            #df_sums = pd.merge(df_sums, conteo_anual, on='year', how='outer')

    
    
    for col in df_analisis.columns:
        if col != 'year':
            # Agrupar por año y calcular la suma de datos anuales
            suma_anual = df_analisis.groupby('year')[col].sum().reset_index()
            
            suma_anual.columns = ['year', f'{col}_sum']
            
            #suma_log.
            promedio_anual = df_analisis.groupby('year')[col].mean().reset_index()
            promedio_anual.columns = ['year', f'{col}_mean']
            promedio_anual[f'{col}_logs'] = suma_anual[f'{col}_sum'].apply(lambda x: np.log(x) if x > 0 else np.nan)
            
            if r_log.empty:
                r_log = promedio_anual[['year', f'{col}_logs']]
            else:
                r_log = pd.merge(r_log, promedio_anual[['year', f'{col}_logs']], on='year', how='outer')
            
            promedio_anual.loc[promedio_anual[f'{col}_mean'] == 0, f'{col}_logs'] = 0

            desviacion_estandar_anual = df_analisis.groupby('year')[col].std().reset_index()
            
            desviacion_estandar_anual.columns = ['year', f'{col}_std']
            
            if resultados_med.empty: # Unir los resultados de suma y promedio en el DataFrame de resultados
                resultados_med = suma_anual
            else:
                resultados_med = pd.merge(resultados_med, suma_anual, on='year', how='outer')
                
            if r_log2.empty:
                r_log2 = promedio_anual[['year', f'{col}_logs']]
            else:
                r_log2 = pd.merge(r_log2, promedio_anual[['year', f'{col}_logs']], on='year', how='outer')
            
            #r_log2 = pd.merge(r_log, promedio_anual[['year', f'{col}_logs']], on='year', how='outer')
            
            resultados_med = pd.merge(resultados_med, promedio_anual[['year', f'{col}_mean', f'{col}_logs']], on='year', how='outer')
            resultados_med = pd.merge(resultados_med, desviacion_estandar_anual, on='year', how='outer')
            

    resultados_med.set_index('year', inplace=True)
    r_log.set_index('year', inplace=True)
    r_log2.set_index('year', inplace=True)
    dd_limites = r_log2.copy()
    
    stats_sum = resultados_med.sum(axis=0).to_frame().T
    stats_cout = resultados_med.count(axis=0).to_frame().T
    stats_mean = resultados_med.mean(axis=0).to_frame().T
    stats_std = resultados_med.std(axis=0).to_frame().T
    
    stats_sum.index = ['Sum']
    stats_cout.index = ['Count']
    stats_mean.index = ['Mean']
    stats_std.index = ['Std']

    # resultados_med2 = pd.concat([resultados_med, stats_sum])

    resultados_med = pd.concat([resultados_med, stats_sum, stats_cout, stats_mean, stats_std])
    
    stats_sum_l = r_log.sum(axis=0).to_frame().T
    stats_cout_l = r_log.count(axis=0).to_frame().T
    stats_mean_l = r_log.mean(axis=0).to_frame().T
    stats_std_l = r_log.std(axis=0).to_frame().T
    stats_lg_l = stats_cout_l.map(lambda x: np.log(x)) # applymap en caso de error
    st_ks =  0.4068 * stats_lg_l + 1.1643
    st_ls = stats_mean_l + stats_std_l * st_ks
    st_li = stats_mean_l - stats_std_l * st_ks
    
    stats_sum_l.index = ['Sum']
    stats_cout_l.index = ['Count']
    stats_mean_l.index = ['Mean']
    stats_std_l.index = ['Std']
    st_ks.index = ['Ks']
    st_ls.index = ['LS']
    st_li.index = ['LI']

    r_log = pd.concat([r_log, stats_sum_l, stats_cout_l, stats_mean_l, stats_std_l, st_ks, st_ls, st_li])
        
    st_ls_compara = st_ls.iloc[0]
    st_li_compara = st_li.iloc[0]
    resultados_comparacion = []
    
    # Iterar sobre las filas de df1 y comparar con la primera fila de df2
    for index, row in dd_limites.iterrows():
        comparacion_s = row < st_ls_compara
        comparacion_i = row > st_li_compara
        resultado_limites = row.where(~(comparacion_i & comparacion_s), "")
        resultados_comparacion.append(resultado_limites)
    
    # Convertir la lista de resultados en un DataFrame
    res_compara_limites = pd.DataFrame(resultados_comparacion, columns=dd_limites.columns)
    res_compara_limites.index.names = ['year']   

    
    resultados_agrupa2 = resultados_agrupa.copy()
    resultados_agrupa2.set_index('year', inplace=True)
    resultados_comparacion = pd.DataFrame(columns=resultados_agrupa2.columns)
    
    for i in range(len(resultados_agrupa2)):
        for j in range(len(resultados_agrupa2.columns)):
            # Obtener el valor en la posición (i, j) de df1 y df2
            valor_df1 = resultados_agrupa2.iloc[i, j]
            valor_df2 = res_compara_limites.iloc[i, j]
            # Comparar los valores y asignar el valor correspondiente de df2 si ambos son mayores que 0 o nulos
            if valor_df1 > 350 and pd.notna(valor_df2):
                valor_nuevo = valor_df2
            else:
                valor_nuevo = ""
            
            # Asignar el nuevo valor al DataFrame de resultados
            resultados_comparacion.at[i, resultados_agrupa2.columns[j]] = valor_nuevo
    resultados_comparacion.index = resultados_agrupa2.index
    res_compara_limites = pd.concat([res_compara_limites, st_ls, st_li])
    
    st_li.columns = resultados_comparacion.columns
    st_ls.columns = resultados_comparacion.columns
    resultados_comparacion = pd.concat([resultados_comparacion, st_ls, st_li])
    
    with pd.ExcelWriter(ruta_completa, engine='openpyxl') as writer:
    # Escribir los DataFrames en las pestañas
        df_pivot.to_excel(writer, sheet_name='Datos diarios')    
        resultados_agrupa.to_excel(writer, sheet_name='Agrupados anuales', index=False)
        resultados_med.to_excel(writer, sheet_name="Estadisticas_anuales", index=True)
        df_sums.to_excel(writer, sheet_name="suma", index=False)
        #r_log.to_excel(writer, sheet_name='LogN', index= True)
        res_compara_limites.to_excel(writer, sheet_name='Límites', index= True)
        resultados_comparacion.to_excel(writer, sheet_name='Líms_VS_añoscompletos', index= True)
        
    print(f"Los DataFrames se han guardado en las pestañas del archivo '{ruta_completa}'")
    
    wb = load_workbook(ruta_completa)

    # def ajustar_tamaño_columnas(worksheet):
    #     for col in worksheet.columns:
    #         max_length = 0
    #         column = col[0].column_letter
    #         for cell in col:
    #             try:
    #                 if len(str(cell.value)) > max_length:
    #                     max_length = len(str(cell.value))
    #             except:
    #                 pass
    #         adjusted_width = (max_length)
    #         worksheet.column_dimensions[column].width = adjusted_width
    
    # Ajustar el tamaño de las columnas en cada pestaña
    # for sheet in wb.sheetnames:
    #     ws = wb[sheet]
    #     ajustar_tamaño_columnas(ws)
    
    def formatear_numeros_dos_decimales(worksheet):
        for row in worksheet.iter_rows(min_row=2):  # Omitir encabezados
            for cell in row:
                if isinstance(cell.value, (float)):
                    cell.number_format = numbers.FORMAT_NUMBER_00
    
    # Ajustar el tamaño de las columnas y formatear números en cada pestaña
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        #ajustar_tamaño_columnas(ws)
        formatear_numeros_dos_decimales(ws)

    wb.save(ruta_completa)
        
except mariadb.Error as e:
    print(f"Error conectando a MariaDB: {e}")
    sys.exit(1)

finally:
    if conn:
        conn.close()
